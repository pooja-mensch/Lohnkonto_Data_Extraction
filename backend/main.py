import sys
import time
import warnings
from copy import copy
from math import floor, ceil

from openpyxl import load_workbook
from pypdf import PdfReader, PasswordType

from misc import data_holder
from misc.classes import Person
from data_extractors.data_extractor_1 import DataExtractor1
from data_extractors.data_extractor_2 import DataExtractor2
from data_extractors.data_extractor_3 import DataExtractor3
from data_extractors.data_extractor_4 import DataExtractor4
from data_extractors.data_extractor_5 import DataExtractor5
from data_extractors.data_extractor_6 import DataExtractor6
from data_extractors.data_extractor_7 import DataExtractor7
from misc.data_holder import *
from meta_detectors.meta_detector_1 import MetaDetector1
from meta_detectors.meta_detector_2 import MetaDetector2
from meta_detectors.meta_detector_3 import MetaDetector3
from meta_detectors.meta_detector_4 import MetaDetector4
from meta_detectors.meta_detector_5 import MetaDetector5
from meta_detectors.meta_detector_6 import MetaDetector6
from meta_detectors.meta_detector_7 import MetaDetector7
from misc.notes_extractor import detect_notes

detectors = { # PyInstaller messes with reflection based access
    MetaDetector1(): DataExtractor1(),
    MetaDetector2(): DataExtractor2(),
    MetaDetector3(): DataExtractor3(),
    MetaDetector4(): DataExtractor4(),
    MetaDetector5(): DataExtractor5(),
    MetaDetector6(): DataExtractor6(),
    MetaDetector7(): DataExtractor7(),
}

people: list[Person] = []

def __push_people(months):
    people.append(Person(
        get_current_meta(),
        copy(months)
    ))

def process_pdf(pdf_file, template, password=None):
    """
    Process a PDF file and generate Excel output.

    Args:
        pdf_file: Path to the PDF file or file-like object
        template: Path to the Excel template file

    Returns:
        tuple: (output_filename, people_count, processing_time)
    """
    reader = PdfReader(pdf_file)
    if reader.is_encrypted:
        if password:
            decrypt_result = reader.decrypt(password)
            if decrypt_result == PasswordType.NOT_DECRYPTED:
                raise ValueError("Incorrect password provided")
        else:
            raise ValueError("PDF is encrypted. Please provide a password.")

    data_holder.set_pages(reader.pages)
    data_holder.set_pages(reader.pages)

    meta_detector = None
    data_extractor = None
    meta: MetaData | None = None

    for detector in detectors.keys():
        meta = detector.detect_meta(reader.pages[0])
        if not meta is None:
            meta_detector = detector
            data_extractor = detectors[detector]
            print(f"[INFO] Erkanntes Format: {detector.__class__.__name__[len('MetaDetector'):]}")
            break

    if meta_detector is None:
        raise ValueError("Dieses Format wird derzeit nicht unterstützt.")


    start = time.time()

    year = meta.year
    client_name = meta.client_name
    print(f"[INFO] Klient: {client_name}")
    print(f"[INFO] Jahr: {year}")

    previous_percent = -1

    for page in reader.pages:
        set_current_page(page.page_number)
        percent = (page.page_number * 20 / len(reader.pages)).__floor__() * 5 # Percentage, rounded to 5% increments
        if percent > previous_percent:
            previous_percent = percent
            print(f"{percent}%")

        meta = meta_detector.detect_meta(page)
        start_page = False
        if not meta is None:
            start_page = True
            if year != meta.year or client_name != meta.client_name:
                print(f"[ERROR] Meta mismatch: {year} != {meta.year} or {client_name} != {meta.client_name}")
                return None

            if not get_current_meta() is None and (meta.surname != get_current_meta().surname or meta.name != get_current_meta().name):
                months = data_extractor.months
                __push_people(months)
                data_extractor.month = -1
                for i in range(0, len(months)):
                    months[i] = {}
            set_current_meta(meta)

        detect_notes(page)
        data_extractor.process_page(page, start_page)

    __push_people(data_extractor.months)
    print("100%")

    with warnings.catch_warnings(action="ignore"):
        workbook = load_workbook(template)
    data = workbook['IST Gehälter']
    project = workbook['Projektübersicht']

    project['B1'] = client_name
    project['B5'] = "01.01." + str(year)
    data['D4'] = ""

    file = 1

    i = 4
    for person in people:
        if i - 4 >= 30: # Too many employees, generate another xlsx file
            workbook.save(f'{client_name} {file}.xlsx')
            with warnings.catch_warnings(action="ignore"):
                workbook = load_workbook(template)
            data = workbook['IST Gehälter']
            project = workbook['Projektübersicht']
            project['B1'] = client_name
            project['B5'] = "01.01." + str(year)
            data['D4'] = ""
            i = 4
            file += 1

        project_column = str(i - 1)
        column = str(i)

        project["E" + project_column] = person.meta.name
        project["F" + project_column] = person.meta.surname
        data["B" + column] = person.meta.name
        data["C" + column] = person.meta.surname
        data["D" + column] = person.meta.weekly_hours
        data["E" + column] = person.meta.first_start
        data["F" + column] = person.meta.start
        data["G" + column] = person.meta.end
        data["H" + column] = ", ".join(person.meta.notes)

        def get_cell_value(month, key):
            value = month[key] if key in month else 0

            if value == "":
                return value
            else:
                return value / 100 # Value is in cents

        month_counter = 0
        for __month in person.months:
            if "BRUTTO" in __month:
                data.cell(row=i, column=month_counter + 9, value=get_cell_value(__month, "BRUTTO"))
                data.cell(row=i, column=month_counter + 35, value=get_cell_value(__month, "RV"))
                data.cell(row=i, column=month_counter + 48, value=get_cell_value(__month, "AV"))
                data.cell(row=i, column=month_counter + 61, value=get_cell_value(__month, "KV"))
                data.cell(row=i, column=month_counter + 74, value=get_cell_value(__month, "PV"))
            month_counter += 1
        i += 1

    if file == 1:
        workbook.save(f'{client_name}.xlsx')
    else:
        workbook.save(f'{client_name} {file}.xlsx')

    diff = ceil(time.time() - start)
    minutes = floor(diff / 60)
    seconds = (diff % 60)
    seconds_string = str(seconds).zfill(2)
    persons_string = "Person wurde" if len(people) == 1 else "Personen wurden"
    print(f"[INFO] {len(people)} {persons_string} in {minutes}:{seconds_string} extrahiert.")

    # Return output filename, count, and time for API usage
    if file == 1:
        output_file = f'{client_name}.xlsx'
    else:
        output_file = f'{client_name} {file}.xlsx'

    return output_file, len(people), diff

def __main():
    """CLI entry point"""
    if len(sys.argv) < 3:
        print("Usage: python main.py <pdf_file> <template_file>")
        return

    pdf_file = sys.argv[1]
    template = sys.argv[2]

    try:
        output_file, people_count, processing_time = process_pdf(pdf_file, template)
        print(f"[INFO] Output saved to: {output_file}")
    except Exception as e:
        print(f"[ERROR] {str(e)}")

if __name__ == '__main__':
    # Use function to be able to return on errors (`exit()` causes problems when converted to .exe)
    __main()